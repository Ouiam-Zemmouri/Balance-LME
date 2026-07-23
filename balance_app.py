import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import glob
import os
import re
import openpyxl

# ===== PASSWORD PROTECTION =====
PASSWORD = "Finance26"

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.set_page_config(page_title="Balance LME", page_icon="⚖️", layout="centered")
    st.markdown("""
    <style>
    :root, .stApp{
      --background-color:#f4f6fb !important;
      --secondary-background-color:#ffffff !important;
      --text-color:#16264a !important;
      --primary-color:#1e3a6d !important;
    }
    .main,[data-testid="stAppViewContainer"]{background:#f4f6fb;}
    input{color:#16264a !important;background:#ffffff !important;}
    </style>
    """, unsafe_allow_html=True)
    st.markdown("<br><br>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1,1.2,1])
    with c2:
        st.markdown("""<div style="text-align:left;margin-bottom:10px;">
          <div style="font-size:1.2rem;font-weight:800;color:#16264a;">⚖️ Balance LME</div>
          <div style="font-size:0.8rem;color:#5b6478;margin-top:2px;">COFICAB Kenitra · COFICAB Maroc</div>
        </div>""", unsafe_allow_html=True)
        password = st.text_input("Password", type="password")
        if st.button("Login", use_container_width=True):
            if password == PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Incorrect password")
    st.stop()

st.set_page_config(page_title="Balance LME", page_icon="⚖️",
                   layout="wide", initial_sidebar_state="expanded")

# ═══════════════════ LIGHT CORPORATE THEME ═══════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* Force light theme at the root — fixes native widgets (multiselect, inputs, sidebar)
   that otherwise inherit Streamlit's dark default regardless of our custom CSS below */
:root, .stApp{
  --background-color:#f4f6fb !important;
  --secondary-background-color:#ffffff !important;
  --text-color:#16264a !important;
  --primary-color:#1e3a6d !important;
}
html,body,[class*="css"]{font-family:'Inter',sans-serif;color:#16264a;}
.main,[data-testid="stAppViewContainer"]{background:#f4f6fb;}
[data-testid="stHeader"]{background:transparent;}
.block-container{padding-top:1.4rem;padding-bottom:2rem;max-width:1400px;}

/* Sidebar — force white background + dark text on every element inside it */
[data-testid="stSidebar"]{background:#ffffff !important;border-right:1px solid #e6e9f2;}
[data-testid="stSidebar"] *{color:#16264a !important;}
[data-testid="stSidebar"] hr{border-color:#e6e9f2 !important;}

/* Safety net: force dark text everywhere in the main content area — inline styles
   (KPI values, badges, etc.) still win since they carry higher CSS specificity */
[data-testid="stAppViewContainer"] *{color:#16264a;}

/* Expanders (native widget headers) were rendering white-on-white — force dark text */
[data-testid="stExpander"] summary,
[data-testid="stExpander"] summary *,
[data-testid="stExpander"] p{color:#16264a !important;}
[data-testid="stExpander"]{background:#ffffff !important;border:1px solid #e9edf5 !important;}
[data-testid="stExpander"] summary,
[data-testid="stExpander"] details,
[data-testid="stExpander"] details[open],
[data-testid="stExpander"] details[open] summary,
[data-testid="stExpander"] summary:hover,
[data-testid="stExpander"] summary:focus,
[data-testid="stExpander"] summary:active{
  background:#ffffff !important;
  background-color:#ffffff !important;
}
[data-testid="stExpanderDetails"]{background:#ffffff !important;}

/* Multiselect / select widgets — force white control + dropdown backgrounds everywhere */
div[data-baseweb="select"] > div{background:#ffffff !important;border-color:#c7d7f2 !important;}
div[data-baseweb="popover"]{background:#ffffff !important;}
ul[role="listbox"]{background:#ffffff !important;}
li[role="option"]{background:#ffffff !important;color:#16264a !important;}
li[role="option"]:hover{background:#eaf0fb !important;}
input{color:#16264a !important;}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif;}
.main,[data-testid="stAppViewContainer"]{background:#f4f6fb;}
[data-testid="stHeader"]{background:transparent;}
.block-container{padding-top:1.4rem;padding-bottom:2rem;max-width:1400px;}

/* Sidebar */
[data-testid="stSidebar"]{background:#ffffff;border-right:1px solid #e6e9f2;}
[data-testid="stSidebar"] .stCaption{color:#8993a8;}

/* Header banner */
.title-banner{
  background:linear-gradient(120deg,#16264a 0%,#1e3a6d 100%);
  border-radius:16px;padding:26px 36px;margin-bottom:22px;
  display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;
  box-shadow:0 8px 28px rgba(20,35,70,0.18);
}
.title-banner h1{
  font-size:1.65rem;font-weight:800;letter-spacing:1px;margin:0;color:#ffffff;
}
.title-banner p{color:#a8c0e8;font-size:0.82rem;margin:4px 0 0 0;letter-spacing:0.5px;}
.badge-strip{display:flex;gap:10px;flex-wrap:wrap;}
.badge{
  background:rgba(255,255,255,0.1);border:1px solid rgba(255,255,255,0.18);
  border-radius:20px;padding:6px 16px;color:#e8eefc;font-size:0.76rem;font-weight:600;
}

/* KPI cards */
.kpi-card{
  background:#ffffff;border-radius:14px;padding:16px 18px;position:relative;overflow:hidden;
  border:1px solid #e9edf5;
  box-shadow:0 2px 10px rgba(20,35,70,0.05);
  height:112px;display:flex;flex-direction:column;justify-content:flex-start;
  transition:all 0.15s ease;
}
.kpi-card:hover{box-shadow:0 6px 20px rgba(20,35,70,0.10);transform:translateY(-1px);}
.kpi-top{display:flex;align-items:center;gap:8px;margin-bottom:10px;}
.kpi-icon{
  width:26px;height:26px;border-radius:8px;display:flex;align-items:center;justify-content:center;
  font-size:0.82rem;flex-shrink:0;
}
.kpi-label{color:#8993a8;font-size:0.66rem;font-weight:700;text-transform:uppercase;letter-spacing:1.1px;}
.kpi-value{color:#16264a;font-size:1.42rem;font-weight:800;line-height:1.1;}
.kpi-sub{color:#a3abbd;font-size:0.7rem;margin-top:3px;font-weight:500;}
.kpi-spark{position:absolute;right:0;bottom:0;width:96px;height:34px;opacity:0.9;}

/* Section headers */
.section-header{
  color:#16264a;font-size:0.92rem;font-weight:700;
  margin:0 0 14px 0;display:flex;align-items:center;gap:8px;
}
.section-sub{color:#8993a8;font-size:0.76rem;margin:-10px 0 14px 0;}

/* Chart / content cards */
div[data-testid="stVerticalBlockBorderWrapper"]{
  background:#ffffff;border-radius:14px;border:1px solid #e9edf5 !important;
  box-shadow:0 2px 10px rgba(20,35,70,0.04);
}

/* Filter pills (multiselect) */
[data-baseweb="tag"]{
  background:#eaf0fb !important;border:1px solid #c7d7f2 !important;
  color:#1e3a6d !important;border-radius:8px !important;font-weight:600 !important;
}
.filter-label{color:#5b6478;font-size:0.68rem;font-weight:700;text-transform:uppercase;
  letter-spacing:1.5px;margin:14px 0 4px 0;}

/* Buttons */
.stButton>button{
  background:#1e3a6d;color:#ffffff;border-radius:9px;border:none;font-weight:600;
}
.stButton>button:hover{background:#16264a;color:#ffffff;}

/* Expander */
.streamlit-expanderHeader{background:#ffffff;border-radius:10px;font-weight:600;color:#16264a;}

/* Dataframe */
.stDataFrame{border-radius:12px;overflow:hidden;border:1px solid #e9edf5;}

/* Gains/Losses dark cards */
.gl-panel{background:#1c1f26;border-radius:16px;padding:22px 20px;}
.gl-panel-title{color:#ffffff;font-size:1rem;font-weight:800;margin-bottom:16px;}
.gl-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:14px;}
.gl-card{background:#2a2e38;border-radius:12px;padding:18px 16px;text-align:center;}
.gl-card-label{color:#9aa0ad;font-size:0.75rem;font-weight:500;margin-bottom:8px;}
.gl-card-value{color:#ffffff;font-size:1.6rem;font-weight:800;line-height:1.15;}
.gl-card-delta{font-size:0.85rem;font-weight:700;margin-top:8px;}

::-webkit-scrollbar{width:6px;height:6px;}
::-webkit-scrollbar-track{background:#f4f6fb;}
::-webkit-scrollbar-thumb{background:#c7d1e3;border-radius:6px;}
</style>
""", unsafe_allow_html=True)

# ── PALETTE (light theme, corporate) ──
NAVY    = "#16264a"   # deep navy — headers, primary text
NAVY_MD = "#1e3a6d"   # medium navy — primary accent / bars
NAVY_LT = "#3d6fc4"   # lighter blue — secondary series
COPPER  = "#c2703d"   # brand copper — COFICAB accent
GOLD    = "#c9932e"   # amber accent
TEAL    = "#0d9488"   # favorable / positive
ROSE    = "#e11d48"   # unfavorable / negative
SLATE   = "#5b6478"   # muted axis / caption text
INK     = "#16264a"   # dark text on white
WHITE   = "#ffffff"

PALETTE = [NAVY_MD, COPPER, TEAL, GOLD, NAVY_LT, "#7c5cbf"]

LAY = dict(
    template="plotly_white",
    paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
    font=dict(color=SLATE, family="Inter", size=12),
    margin=dict(t=30, b=40, l=55, r=20),
    legend=dict(bgcolor="rgba(255,255,255,0.95)", bordercolor="#e9edf5",
                borderwidth=1, font=dict(color=INK, size=11)),
    xaxis=dict(gridcolor="#f0f2f8", zeroline=False, tickfont=dict(color=SLATE, size=11)),
    yaxis=dict(gridcolor="#f0f2f8", zeroline=False, tickfont=dict(color=SLATE, size=11)),
    hoverlabel=dict(bgcolor="#ffffff", bordercolor="#e9edf5", font=dict(color=INK, size=12)),
)

import io
import base64
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

def sparkline_b64(values, color):
    """Tiny transparent area-sparkline as a base64 PNG, for embedding inside a KPI card."""
    if values is None or len(values) < 2:
        return None
    fig, ax = plt.subplots(figsize=(2.0, 0.5), dpi=150)
    ax.plot(values, color=color, linewidth=2)
    ax.fill_between(range(len(values)), values, min(values), color=color, alpha=0.12)
    ax.axis("off")
    fig.subplots_adjust(left=0, right=1, top=1, bottom=0)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", transparent=True)
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode()

def kpi(col, icon, label, val, color=NAVY_MD, sub=None, spark=None):
    sub_html = f'<div class="kpi-sub">{sub}</div>' if sub else ""
    spark_html = ""
    if spark:
        b64 = sparkline_b64(spark, color)
        if b64:
            spark_html = f'<img class="kpi-spark" src="data:image/png;base64,{b64}"/>'
    col.markdown(f"""<div class="kpi-card">
      <div class="kpi-top">
        <div class="kpi-icon" style="background:{color}1a;color:{color};">{icon}</div>
        <div class="kpi-label">{label}</div>
      </div>
      <div class="kpi-value">{val}</div>
      {sub_html}
      {spark_html}
      </div>""", unsafe_allow_html=True)

def fmt_compact(value):
    """Format a euro amount compactly, e.g. 1990000 -> '1.99 M', 374400 -> '374.4 K'."""
    sign = "-" if value < 0 else ""
    v = abs(value)
    if v >= 1_000_000:
        return f"{sign}{v/1_000_000:.2f} M"
    if v >= 1_000:
        return f"{sign}{v/1_000:.1f} K"
    return f"{sign}{v:,.0f}"

def sec(icon, title, sub=None):
    st.markdown(f'<div class="section-header">{icon}&nbsp; {title}</div>', unsafe_allow_html=True)
    if sub:
        st.markdown(f'<div class="section-sub">{sub}</div>', unsafe_allow_html=True)

def alay(fig, **kw):
    fig.update_layout(**{**LAY, **kw}); return fig

def entity_color_map(entities):
    return {e: PALETTE[i % len(PALETTE)] for i, e in enumerate(entities)}

# ── LME BALANCE — PARSING ──
ENTITY_CODE_MAP = {
    "KT": "COFICAB Kenitra", "KEN": "COFICAB Kenitra",
    "MR": "COFICAB Maroc", "MA": "COFICAB Maroc",
    "INTL": "COFICAB International", "INT": "COFICAB International",
}
MONTH_NUM_MAP = {"01":"Jan","02":"Feb","03":"Mar","04":"Apr","05":"May","06":"Jun",
                  "07":"Jul","08":"Aug","09":"Sep","10":"Oct","11":"Nov","12":"Dec"}
BALANCE_FOLDER = "balance_files"

def parse_lme_balance_file(path):
    """Parse one monthly 'LME balance calculation under FIFO method' Excel file."""
    fname = os.path.basename(path)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return None, f"❌ {fname}: unable to read file ({e})"

    ws = wb[wb.sheetnames[0]]
    header_row = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 2).value).strip() == "LME Projects":
            header_row = r
            break
    if header_row is None:
        return None, f"❌ {fname}: could not locate the 'LME Projects' balance table."

    rows = []
    r = header_row + 1
    while r <= ws.max_row:
        proj = ws.cell(r, 2).value
        if proj is None or str(proj).strip() == "":
            break
        proj = str(proj).strip()
        rows.append({
            "Fixation":        proj,
            "Qty_Sold_T":      ws.cell(r, 3).value,
            "LME_Sales":       ws.cell(r, 4).value,
            "Sales_Value":     ws.cell(r, 5).value,
            "Qty_Stock_T":     ws.cell(r, 6).value,
            "LME_Stock":       ws.cell(r, 7).value,
            "Stock_Value":     ws.cell(r, 8).value,
            "Qty_Purchase_T":  ws.cell(r, 9).value,
            "LME_Purchase":    ws.cell(r, 10).value,
            "Purchase_Value":  ws.cell(r, 11).value,
            "Needs_Exceed_T":  ws.cell(r, 12).value,
            "Allocated_QTE":   ws.cell(r, 13).value,
            "LME_Realloc":     ws.cell(r, 14).value,
            "Realloc_Value":   ws.cell(r, 15).value,
            "Last_QTY_T":      ws.cell(r, 16).value,
            "LME_Final":       ws.cell(r, 17).value,
            "Final_Value":     ws.cell(r, 18).value,
            "LME_Balance_Eur": ws.cell(r, 19).value,
        })
        if proj.upper() == "TOTAL":
            break
        r += 1

    if not rows:
        return None, f"❌ {fname}: no fixation rows found under the balance table."

    df_bal = pd.DataFrame(rows)
    num_cols = [c for c in df_bal.columns if c != "Fixation"]
    df_bal[num_cols] = df_bal[num_cols].apply(pd.to_numeric, errors="coerce")

    m = re.search(r'COF[\s_-]*([A-Za-z]+).*?(\d{1,2})[\s._-](\d{4})', fname)
    if m:
        entity = ENTITY_CODE_MAP.get(m.group(1).upper(), f"COFICAB {m.group(1).upper()}")
        month_num, year = m.group(2).zfill(2), m.group(3)
        month_label = f"{MONTH_NUM_MAP.get(month_num, month_num)} {year}"
        month_key = f"{year}-{month_num}"
    else:
        entity, month_label, month_key = "Unknown", fname, fname

    df_bal.insert(0, "Entity", entity)
    df_bal.insert(1, "Month", month_label)
    df_bal.insert(2, "MonthKey", month_key)
    df_bal.insert(3, "SourceFile", fname)
    return df_bal, None

@st.cache_data(ttl=3600)
def load_all_balance_files():
    """Auto-load every .xlsx bundled in the balance_files/ folder of the repo."""
    files = sorted(
        f for f in glob.glob(os.path.join(BALANCE_FOLDER, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")
    )
    dfs, errs = [], []
    for path in files:
        df, err = parse_lme_balance_file(path)
        (errs if err else dfs).append(err if err else df)
    if not dfs:
        return pd.DataFrame(), errs
    return pd.concat(dfs, ignore_index=True), errs

def generate_balance_insights(view_fix, view_tot):
    insights = []
    if view_fix.empty or view_tot.empty:
        return ["Not enough data to generate an analysis for the current selection."]

    vf = view_fix.dropna(subset=["LME_Sales"])
    if not vf.empty:
        best = vf.loc[vf["LME_Sales"].idxmin()]
        worst = vf.loc[vf["LME_Sales"].idxmax()]
        if best["Fixation"] != worst["Fixation"]:
            insights.append(
                f"The most favorable sales fixation is **{best['Fixation']}** at "
                f"**{best['LME_Sales']:.4f} €/kg**, while **{worst['Fixation']}** carries the "
                f"highest sales price at **{worst['LME_Sales']:.4f} €/kg**."
            )

    vs = view_fix.dropna(subset=["Qty_Stock_T"])
    vs = vs[vs["Qty_Stock_T"] > 0]
    if not vs.empty:
        top_stock = vs.loc[vs["Qty_Stock_T"].idxmax()]
        insights.append(
            f"**{top_stock['Fixation']}** carries the largest stock position at "
            f"**{top_stock['Qty_Stock_T']:.1f} T**, making it the main buffer in the FIFO valuation chain."
        )

    ne = view_fix.dropna(subset=["Needs_Exceed_T"])
    deficits = ne[ne["Needs_Exceed_T"] > 0.5]
    surplus  = ne[ne["Needs_Exceed_T"] < -0.5]
    if not deficits.empty:
        d = deficits.iloc[0]
        if not surplus.empty:
            s = surplus.iloc[0]
            insights.append(
                f"Sold quantities on **{d['Fixation']}** exceeded available stock and purchases by "
                f"**{d['Needs_Exceed_T']:.1f} T**; under FIFO this shortfall is reallocated from the "
                f"surplus fixation **{s['Fixation']}**."
            )
        else:
            insights.append(
                f"Sold quantities on **{d['Fixation']}** exceeded available stock and purchases by "
                f"**{d['Needs_Exceed_T']:.1f} T**, requiring reallocation from the next fixation in the FIFO sequence."
            )

    tot_balance = view_tot["LME_Balance_Eur"].sum()
    direction = "favorable" if tot_balance >= 0 else "unfavorable"
    insights.append(
        f"The overall LME balance is **{direction}**, at **€{tot_balance:,.0f}** — sales were valued "
        f"{'above' if tot_balance >= 0 else 'below'} the FIFO cost of stock and purchases consumed."
    )

    vb = view_fix.dropna(subset=["LME_Balance_Eur"])
    if not vb.empty:
        top_c = vb.reindex(vb["LME_Balance_Eur"].abs().sort_values(ascending=False).index).iloc[0]
        sign = "gain" if top_c["LME_Balance_Eur"] >= 0 else "loss"
        insights.append(
            f"**{top_c['Fixation']}** is the largest single contributor to this result, driving a "
            f"**{sign} of €{abs(top_c['LME_Balance_Eur']):,.0f}**."
        )
    return insights
# ── SIDEBAR ──
with st.sidebar:
    st.markdown("""<div style="padding:4px 0 10px 0;">
      <div style="font-size:1.05rem;font-weight:800;color:#16264a;">⚖️ Balance LME</div>
      <div style="font-size:0.78rem;color:#5b6478;margin-top:2px;">COFICAB Kenitra · COFICAB Maroc</div>
    </div>""", unsafe_allow_html=True)
    if st.button("🔄 Refresh Data", use_container_width=True):
        st.cache_data.clear()
        st.rerun()
    st.markdown("---")

bal_all, errors = load_all_balance_files()
for e in errors:
    st.warning(e)

if bal_all.empty:
    st.markdown("""<div class="title-banner">
      <div><h1>⚖️ Balance LME</h1><p>FIFO Method · Sales vs Stock &amp; Purchase Valuation</p></div>
    </div>""", unsafe_allow_html=True)
    st.info(
        f"No balance files found yet. Add your monthly `.xlsx` files "
        f"(e.g. `COF_KT_-_LME_balance_06_2026.xlsx`) to the **`{BALANCE_FOLDER}/`** "
        f"folder in this repo, commit, then click 🔄 Refresh Data in the sidebar."
    )
    st.stop()

bal_all["Group"] = bal_all["Entity"] + " · " + bal_all["Month"]
ENT_COLOR = entity_color_map(sorted(bal_all["Entity"].unique()))

# ── FILTERS (sidebar) ──
with st.sidebar:
    st.markdown('<p class="filter-label">Entity</p>', unsafe_allow_html=True)
    ent_opts = sorted(bal_all["Entity"].unique())
    sel_e = st.multiselect("", ent_opts, default=ent_opts, key="bal_ent", label_visibility="collapsed")

    st.markdown('<p class="filter-label">Month</p>', unsafe_allow_html=True)
    month_map = bal_all[["MonthKey","Month"]].drop_duplicates().sort_values("MonthKey")
    sel_m = st.multiselect("", month_map["Month"].tolist(), default=month_map["Month"].tolist(),
                            key="bal_month", label_visibility="collapsed")

    st.markdown("---")
    st.caption(f"📁 {bal_all['SourceFile'].nunique()} file(s) loaded")
    st.caption(f"🏭 {bal_all['Entity'].nunique()} entit{'y' if bal_all['Entity'].nunique()==1 else 'ies'} · "
               f"{bal_all['MonthKey'].nunique()} month(s)")

view = bal_all[bal_all["Entity"].isin(sel_e) & bal_all["Month"].isin(sel_m)].copy()
view_fix = view[view["Fixation"].str.upper() != "TOTAL"].copy()
view_tot = view[view["Fixation"].str.upper() == "TOTAL"].copy()

# ── HEADER BANNER ──
st.markdown(f"""<div class="title-banner">
  <div>
    <h1>⚖️ Balance LME</h1>
    <p>FIFO Method · Sales vs Stock &amp; Purchase Valuation</p>
  </div>
  <div class="badge-strip">
    <div class="badge">🏭 {', '.join(sel_e) if len(sel_e)<=2 else f'{len(sel_e)} entities'}</div>
    <div class="badge">📅 {', '.join(sel_m) if len(sel_m)<=3 else f'{len(sel_m)} months'}</div>
  </div>
</div>""", unsafe_allow_html=True)

if view.empty:
    st.warning("⚠️ No data matches the selected Entity / Month filters.")
    st.stop()

groups = sorted(view["Group"].unique())
month_order = month_map[month_map["Month"].isin(sel_m)]["Month"].tolist()

# ══════════════════════ KPI ROW ══════════════════════
tot_sales   = view_tot["Sales_Value"].sum()
tot_final   = view_tot["Final_Value"].sum()
tot_balance = view_tot["LME_Balance_Eur"].sum()
tot_qty     = view_tot["Qty_Sold_T"].sum()
bal_per_t   = tot_balance / tot_qty if tot_qty else 0
bal_color   = TEAL if tot_balance >= 0 else ROSE
n_fav       = (view_tot["LME_Balance_Eur"] >= 0).sum()
n_tot       = len(view_tot)

monthly = view_tot.groupby("MonthKey")[["Sales_Value","Final_Value","LME_Balance_Eur"]].sum().sort_index()
spark_sales   = monthly["Sales_Value"].tolist()   if len(monthly) > 1 else None
spark_final   = monthly["Final_Value"].tolist()   if len(monthly) > 1 else None
spark_balance = monthly["LME_Balance_Eur"].tolist() if len(monthly) > 1 else None

k1,k2,k3,k4,k5 = st.columns(5)
kpi(k1,"💶","Sales Valuation",        f"€{tot_sales:,.0f}",   NAVY_MD, "Total sold, valorized", spark_sales)
kpi(k2,"📦","Stock + Purchase Value", f"€{tot_final:,.0f}",   NAVY_LT, "FIFO cost basis", spark_final)
kpi(k3,"⚖️","Net LME Balance",        f"€{tot_balance:,.0f}", bal_color, "Favorable" if tot_balance>=0 else "Unfavorable", spark_balance)
kpi(k4,"📏","Balance per Ton",        f"€{bal_per_t:,.1f}/T", GOLD, f"on {tot_qty:,.0f} T sold")
kpi(k5,"✅","Favorable Periods",      f"{n_fav} / {n_tot}",   TEAL if n_fav==n_tot else ROSE, "entity × month")


st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

# ══════════════════════ ROW A — TREND + SPLIT ══════════════════════
rowA1, rowA2 = st.columns([2,1])

with rowA1:
    with st.container(border=True):
        sec("📅","Net LME Balance — Trend", "Monthly evolution by entity")
        trend = (view_tot.groupby(["Entity","MonthKey","Month"])["LME_Balance_Eur"]
                 .sum().reset_index().sort_values("MonthKey"))
        if trend["MonthKey"].nunique() > 1:
            figA1 = px.line(trend, x="Month", y="LME_Balance_Eur", color="Entity",
                             markers=True, category_orders={"Month": month_order},
                             color_discrete_map=ENT_COLOR)
            figA1.update_traces(line=dict(width=3), marker=dict(size=9))
            figA1.add_hline(y=0, line_dash="dot", line_color="#dde3f0")
            alay(figA1, showlegend=len(sel_e) > 1,
                 yaxis=dict(title="LME Balance (€)"), xaxis=dict(title=""))
            st.plotly_chart(figA1, use_container_width=True, theme=None)
        else:
            figA1 = px.bar(view_tot, x="Entity", y="LME_Balance_Eur", color="Entity",
                            color_discrete_map=ENT_COLOR, text_auto=",.0f")
            figA1.add_hline(y=0, line_dash="dot", line_color="#dde3f0")
            alay(figA1, showlegend=False, yaxis=dict(title="LME Balance (€)"), xaxis=dict(title=""))
            st.plotly_chart(figA1, use_container_width=True, theme=None)
            st.caption("Add more monthly files to unlock the trend view.")

with rowA2:
    with st.container(border=True):
        sec("🥯","Valuation Split", "Sales vs cost basis")
        donut_df = pd.DataFrame({"Component": ["Sales", "Stock + Purchase"], "Value": [tot_sales, tot_final]})
        figA2 = go.Figure(go.Pie(
            labels=donut_df["Component"], values=donut_df["Value"], hole=0.62,
            marker=dict(colors=[COPPER, NAVY_LT], line=dict(color="#ffffff", width=3)),
            textinfo="percent", textfont=dict(color="#ffffff", size=12)
        ))
        alay(figA2, showlegend=True,
             legend=dict(orientation="h", yanchor="bottom", y=-0.15, xanchor="center", x=0.5),
             annotations=[dict(text=f"€{tot_sales - tot_final:+,.0f}", x=0.5, y=0.5,
                                font=dict(size=15, color=bal_color, family="Inter"), showarrow=False)])
        st.plotly_chart(figA2, use_container_width=True, theme=None)

# ══════════════════════ ROW B — FIXATION BREAKDOWN ══════════════════════
rowB1, rowB2 = st.columns(2)

with rowB1:
    with st.container(border=True):
        sec("📈","LME Balance by Fixation", "Positive = favorable to COFICAB")
        figB1 = px.bar(view_fix, x="Fixation", y="LME_Balance_Eur", color="Group",
                        barmode="group", text_auto=",.0f", color_discrete_sequence=PALETTE)
        figB1.add_hline(y=0, line_color="#dde3f0")
        figB1.update_traces(textfont=dict(size=10, color=INK), textposition="outside")
        alay(figB1, showlegend=len(groups) > 1,
             yaxis=dict(title="LME Balance (€)"), xaxis=dict(title=""),
             legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        st.plotly_chart(figB1, use_container_width=True, theme=None)

with rowB2:
    with st.container(border=True):
        sec("💰","Sales vs Valuation", "By fixation, aggregated across selection")
        melted = view_fix.groupby("Fixation")[["Sales_Value","Final_Value"]].sum().reset_index()
        figB2 = go.Figure()
        figB2.add_trace(go.Bar(name="Sales", y=melted["Fixation"], x=melted["Sales_Value"],
                                orientation="h", marker_color=COPPER))
        figB2.add_trace(go.Bar(name="Stock+Purchase", y=melted["Fixation"], x=melted["Final_Value"],
                                orientation="h", marker_color=NAVY_LT))
        alay(figB2, barmode="group", xaxis=dict(title="Value (€)"), yaxis=dict(title=""),
             legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        st.plotly_chart(figB2, use_container_width=True, theme=None)

# ══════════════════════ FIXATION DETAIL (main focus) ══════════════════════
with st.container(border=True):
    sec("🔍","Fixation Detail", "Full breakdown per fixation, aggregated across the current selection")

    fix_summary = view_fix.groupby("Fixation").agg(
        Qty_Sold_T=("Qty_Sold_T","sum"),
        Qty_Stock_T=("Qty_Stock_T","sum"),
        Qty_Purchase_T=("Qty_Purchase_T","sum"),
        Needs_Exceed_T=("Needs_Exceed_T","sum"),
        Sales_Value=("Sales_Value","sum"),
        Final_Value=("Final_Value","sum"),
        LME_Balance_Eur=("LME_Balance_Eur","sum"),
    ).reset_index()
    fix_summary["Avg_Sales_LME"] = fix_summary["Sales_Value"] / fix_summary["Qty_Sold_T"].replace(0, pd.NA)
    fix_summary["Balance_per_T"] = fix_summary["LME_Balance_Eur"] / fix_summary["Qty_Sold_T"].replace(0, pd.NA)
    fix_summary = fix_summary.sort_values("Fixation")

    disp_fs = fix_summary.rename(columns={
        "Fixation":"Fixation","Qty_Sold_T":"Qty Sold (T)","Qty_Stock_T":"Qty Stock (T)",
        "Qty_Purchase_T":"Qty Purchase (T)","Needs_Exceed_T":"Needs(+)/Exceed(-) (T)",
        "Avg_Sales_LME":"Avg Sales LME (€/kg)","Sales_Value":"Sales Value (€)",
        "Final_Value":"Stock+Purchase Value (€)","LME_Balance_Eur":"LME Balance (€)",
        "Balance_per_T":"Balance per Ton (€/T)"
    })
    disp_fs = disp_fs[["Fixation","Qty Sold (T)","Qty Stock (T)","Qty Purchase (T)",
                        "Needs(+)/Exceed(-) (T)","Avg Sales LME (€/kg)","Sales Value (€)",
                        "Stock+Purchase Value (€)","LME Balance (€)","Balance per Ton (€/T)"]]

    fmt_fs = {
        "Qty Sold (T)":"{:,.1f}", "Qty Stock (T)":"{:,.1f}", "Qty Purchase (T)":"{:,.1f}",
        "Needs(+)/Exceed(-) (T)":"{:,.1f}", "Avg Sales LME (€/kg)":"{:.4f}",
        "Sales Value (€)":"€{:,.0f}", "Stock+Purchase Value (€)":"€{:,.0f}",
        "LME Balance (€)":"€{:,.0f}", "Balance per Ton (€/T)":"€{:,.1f}",
    }
    st.dataframe(
        disp_fs.style.format(fmt_fs)
            .set_properties(**{"background-color":"#ffffff","color":INK})
            .map(lambda v:"color:#0d9488;font-weight:700" if isinstance(v,(int,float)) and v>0
                 else ("color:#e11d48;font-weight:700" if isinstance(v,(int,float)) and v<0 else ""),
                 subset=["LME Balance (€)","Balance per Ton (€/T)"]),
        use_container_width=True, hide_index=True, height=38*len(disp_fs)+40
    )

    fix_trend = (view_fix.groupby(["Fixation","MonthKey","Month"])["LME_Balance_Eur"]
                 .sum().reset_index().sort_values("MonthKey"))
    if fix_trend["MonthKey"].nunique() > 1:
        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
        figFT = px.line(fix_trend, x="Month", y="LME_Balance_Eur", color="Fixation",
                        markers=True, category_orders={"Month": month_order},
                        color_discrete_sequence=PALETTE)
        figFT.update_traces(line=dict(width=3), marker=dict(size=8))
        figFT.add_hline(y=0, line_dash="dot", line_color="#dde3f0")
        alay(figFT, title="LME Balance (€) — Trend per Fixation",
             yaxis=dict(title="LME Balance (€)"), xaxis=dict(title=""),
             legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        st.plotly_chart(figFT, use_container_width=True, theme=None)
    else:
        st.caption("Add more monthly files to unlock the per-fixation trend view.")

# ══════════════════════ ROW C — ENTITY x MONTH COMPARISON ══════════════════════
if len(sel_e) > 1 or len(sel_m) > 1:
    with st.container(border=True):
        sec("🌍","Balance by Entity & Month", "Side-by-side comparison")
        figC = px.bar(view_tot, x="Month", y="LME_Balance_Eur", color="Entity",
                      barmode="group", text_auto=",.0f",
                      category_orders={"Month": month_order}, color_discrete_map=ENT_COLOR)
        figC.add_hline(y=0, line_color="#dde3f0")
        figC.update_traces(textfont=dict(size=10, color=INK), textposition="outside")
        alay(figC, showlegend=True, yaxis=dict(title="LME Balance (€)"), xaxis=dict(title=""),
             legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        st.plotly_chart(figC, use_container_width=True, theme=None)

# ══════════════════════ GAINS / LOSSES ══════════════════════
gl_cards = []
for g in groups:
    sub_tot = view_tot[view_tot["Group"] == g]
    g_bal = sub_tot["LME_Balance_Eur"].sum()
    g_qty = sub_tot["Qty_Sold_T"].sum()
    g_per_t = g_bal / g_qty if g_qty else 0
    is_gain = g_bal >= 0
    arrow = "▲" if is_gain else "▼"
    delta_color = "#3ddc97" if is_gain else "#ff6b6b"
    gl_cards.append(f"""
      <div class="gl-card">
        <div class="gl-card-label">{g}</div>
        <div class="gl-card-value">€{fmt_compact(g_bal)}</div>
        <div class="gl-card-delta" style="color:{delta_color};">{arrow} €{fmt_compact(g_per_t)}/T</div>
      </div>""")

st.markdown(f"""<div class="gl-panel">
  <div class="gl-panel-title">💹 Gains / Losses — LME Balance Result</div>
  <div class="gl-grid">{''.join(gl_cards)}</div>
</div>""", unsafe_allow_html=True)
st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

fix_gl_cards = []
fix_agg = view_fix.groupby("Fixation")[["LME_Balance_Eur","Qty_Sold_T"]].sum().reset_index()
for _, row in fix_agg.iterrows():
    f_bal = row["LME_Balance_Eur"]
    f_qty = row["Qty_Sold_T"]
    f_per_t = f_bal / f_qty if f_qty else 0
    is_gain = f_bal >= 0
    arrow = "▲" if is_gain else "▼"
    delta_color = "#3ddc97" if is_gain else "#ff6b6b"
    fix_gl_cards.append(f"""
      <div class="gl-card">
        <div class="gl-card-label">{row['Fixation']}</div>
        <div class="gl-card-value">€{fmt_compact(f_bal)}</div>
        <div class="gl-card-delta" style="color:{delta_color};">{arrow} €{fmt_compact(f_per_t)}/T</div>
      </div>""")

st.markdown(f"""<div class="gl-panel">
  <div class="gl-panel-title">💹 Gains / Losses — by Fixation</div>
  <div class="gl-grid">{''.join(fix_gl_cards)}</div>
</div>""", unsafe_allow_html=True)
st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# ══════════════════════ INSIGHTS ══════════════════════
with st.container(border=True):
    sec("🧠","Result Interpretation", "Auto-generated from the current selection")
    if len(groups) == 1:
        for line in generate_balance_insights(view_fix, view_tot):
            st.markdown(f"- {line}")
    else:
        for g in groups:
            with st.expander(f"📌 {g}", expanded=False):
                sub_fix = view_fix[view_fix["Group"] == g]
                sub_tot = view_tot[view_tot["Group"] == g]
                for line in generate_balance_insights(sub_fix, sub_tot):
                    st.markdown(f"- {line}")

# ══════════════════════ FULL DATA TABLE ══════════════════════
with st.expander("📋 Full LME Balance Table", expanded=False):
    disp_cols = ["Entity","Month","Fixation","Qty_Sold_T","LME_Sales","Sales_Value",
                 "Qty_Stock_T","LME_Stock","Stock_Value","Qty_Purchase_T","LME_Purchase",
                 "Purchase_Value","Needs_Exceed_T","Last_QTY_T","LME_Final","Final_Value",
                 "LME_Balance_Eur"]
    disp = view[disp_cols].reset_index(drop=True)
    disp.columns = ["Entity","Month","Fixation","Qty Sold (T)","LME Sales (€/kg)","Sales Value (€)",
                    "Qty Stock (T)","LME Stock (€/kg)","Stock Value (€)","Qty Purchase (T)",
                    "LME Purchase (€/kg)","Purchase Value (€)","Needs(+)/Exceed(-) (T)",
                    "Final Qty (T)","LME Final (€/kg)","Final Value (€)","LME Balance (€)"]

    qty_cols = ["Qty Sold (T)","Qty Stock (T)","Qty Purchase (T)","Needs(+)/Exceed(-) (T)","Final Qty (T)"]
    lme_cols = ["LME Sales (€/kg)","LME Stock (€/kg)","LME Purchase (€/kg)","LME Final (€/kg)"]
    eur_cols = ["Sales Value (€)","Stock Value (€)","Purchase Value (€)","Final Value (€)","LME Balance (€)"]
    fmt = {c:"{:,.2f}" for c in qty_cols}
    fmt.update({c:"{:.4f}" for c in lme_cols})
    fmt.update({c:"€{:,.0f}" for c in eur_cols})

    st.dataframe(
        disp.style.format(fmt)
            .set_properties(**{"background-color":"#ffffff","color":INK})
            .map(lambda v:"color:#0d9488;font-weight:700" if isinstance(v,(int,float)) and v>0
                 else ("color:#e11d48;font-weight:700" if isinstance(v,(int,float)) and v<0 else ""),
                 subset=["LME Balance (€)"])
            .map(lambda v:"font-weight:700;color:#c2703d" if str(v).strip().upper()=="TOTAL" else "",
                 subset=["Fixation"]),
        use_container_width=True, hide_index=True, height=380
    )

    st.download_button(
        "⬇️ Download consolidated balance (CSV)",
        data=view[disp_cols].to_csv(index=False).encode("utf-8"),
        file_name="lme_balance_consolidated.csv", mime="text/csv"
    )

st.markdown(f"""<div style="text-align:center;color:#a3abbd;font-size:0.72rem;
  margin-top:40px;padding:16px;border-top:1px solid #e9edf5;">
  Balance LME &nbsp;·&nbsp; COFICAB Kenitra &amp; COFICAB Maroc
</div>""", unsafe_allow_html=True)
